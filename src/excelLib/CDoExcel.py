import time
import openpyxl
from openpyxl.styles import PatternFill

from src.const._const import _const
from src.bean.TaskBean import TaskBean

#import jpype
#import mpxj

class CDoExcel(object):

    _const.PROJECT = 1
    _const.ISSUE_TYPE = 2
    _const.SUMMARY = 3
    _const.TEAMDEVELOPMENT = 4
    _const.FUNCTIONALITY = 5
    _const.ASSIGNEE = 6
    _const.TEAMTEST = 7
    _const.BUGFOUNDBY = 8
    _const.DESCRIPTION = 9
    _const.REPRODUCIBILITY = 10
    _const.SEVERITY = 11
    _const.MANUFACTURERVARIANT = 12
    _const.FOUNDINSW = 13
    _const.FOUNDINHW = 14
    _const.LABELS = 15
    _const.KEY = 16
    _const.COMMENTS = 17
    _const.OSS_CVEID = 14
    _const.OSS_INFLUENCE = 25
    _const.OSS_JIRAURL = 26
    _const.OSS_JIRASTATUS = 27
    _const.OSS_AffectedSW = 28
    _const.OSS_COMMENTS = 29
    _const.OSS_COMPONENT = 4

    __m_workBook = None
    __m_sheet = None
    __m_maxRow = None
    __m_filePath = None

    def __init__(self, _excelFilePath):
        ''' mpp文件读取好用
        jpype.startJVM()
        from net.sf.mpxj.reader import UniversalProjectReader
        project = UniversalProjectReader().read(_excelFilePath)
        
        print("Tasks")
        for task in project.getTasks():
          print(task.getID().toString() + "\t" + task.getName())
          if task.getID() != 0:
            print("\t" + str(task.getStart().getDate()) + "\t" + str(task.getText(16)) + "\t" + str(task.getMarked()))
          if task.getParentTask() is not None:
            print(task.getParentTask())
        
        jpype.shutdownJVM()     
        '''

        # excel
        self.__m_filePath = _excelFilePath
        self.__m_workBook = openpyxl.load_workbook(_excelFilePath)
        
        
        #print('sheet names are ' + str(self.__m_workBook.sheetnames))

        #for name in self.__m_workBook.sheetnames:
        #    print('================')
        #    rowMax = self.__m_workBook[name].max_row
        #    colMax = self.__m_workBook[name].max_column
        #    print('row = ' + str(rowMax) + ', col = ' + str(colMax))        
        #    print('A1 = ' + self.__m_workBook[name].cell(1,1).value)

    def getWorkBook(self):
        return self.__m_workBook

    def setSheetName(self, _name):
        self.__m_sheet = None
        self.__m_maxRow = None
        if self.__m_workBook is not None:
            self.__m_sheet = self.__m_workBook[_name]
            if self.__m_sheet is not None:
                self.__m_maxRow = self.__m_sheet.max_row
    
    def getMaxrow(self):
        return self.__m_maxRow

    def getDataDict(self, _index, _jiarHndl):
        tb = TaskBean()

        _project = self.__m_sheet.cell(_index,_const.PROJECT).value
        tb.setProject(_project)
        _issuetype = self.__m_sheet.cell(_index,_const.ISSUE_TYPE).value
        tb.setIssuetype(_issuetype)
        _summary = self.__m_sheet.cell(_index,_const.SUMMARY).value
        tb.setSummary(_summary)
        _teamdevelopment= str(self.__m_sheet.cell(_index,_const.TEAMDEVELOPMENT).value)
        tb.setteamdevelopment(_teamdevelopment)
        _functionality = self.__m_sheet.cell(_index,_const.FUNCTIONALITY).value
        tb.setFunctionality(_functionality)
        _assignee = self.__m_sheet.cell(_index,_const.ASSIGNEE).value
        tb.setAssignee(_assignee)
        _teamtest = self.__m_sheet.cell(_index,_const.TEAMTEST).value
        tb.setteamtest(_teamtest)
        _buyfoundby = self.__m_sheet.cell(_index, _const.BUGFOUNDBY).value
        tb.setbugfoundby(_buyfoundby)
        _description = self.__m_sheet.cell(_index, _const.DESCRIPTION).value
        tb.setDescription(_description)
        _reproducibility= self.__m_sheet.cell(_index, _const.REPRODUCIBILITY).value
        tb.setreproducibility(_reproducibility)
        _severity = self.__m_sheet.cell(_index, _const.SEVERITY).value
        tb.setSeverity(_severity)
        _manufacturervariant = self.__m_sheet.cell(_index, _const.MANUFACTURERVARIANT).value
        tb.setManufacturervariant(_manufacturervariant)
        _foundinsw = self.__m_sheet.cell(_index, _const.FOUNDINSW).value
        tb.setfoundinSW(_foundinsw)
        _foundinhw = self.__m_sheet.cell(_index, _const.FOUNDINHW).value
        tb.setfoundinHW(_foundinhw)
        _labels = self.__m_sheet.cell(_index,_const.LABELS).value
        label_list = []
        # print(_labels)
        while _labels is not None:
            index = _labels.find(' ')
            if index > 0:
                label = _labels[:index]
                label_list.append(label)
                _labels = _labels[(index+1):]
            else:
                if len(_labels.strip()) > 0:
                    label_list.append(_labels)
                break
        print(label_list)
        tb.setLabels(label_list)

        return tb.getDataDict(_jiarHndl)

    def setJIRA2OSSData(self, _index, _searchlist):
        oss_cveid = self.__m_sheet.cell(_index, _const.OSS_CVEID).value
        oss_component = self.__m_sheet.cell(_index, _const.OSS_COMPONENT).value
        for searchinfo in _searchlist:
            cveid = searchinfo.get('cveid')
            component = searchinfo.get('component')
            status = searchinfo.get('status')
            print(oss_component, component)
            if cveid == oss_cveid and oss_component in component:
                self.__m_sheet.cell(_index, _const.OSS_COMMENTS).value = searchinfo.get('comments')
                self.__m_sheet.cell(_index, _const.OSS_JIRAURL).value = "https://jira.jnd.joynext.com/browse/" + searchinfo.get('jiranum')
                self.__m_sheet.cell(_index, _const.OSS_JIRASTATUS).value = status
                self.__m_sheet.cell(_index, _const.OSS_AffectedSW).value = searchinfo.get('AffectedSW')
                if searchinfo.get('influence') == 'Y' and  ( self.__m_sheet.cell(_index, _const.OSS_INFLUENCE).value is None or  self.__m_sheet.cell(_index, _const.OSS_INFLUENCE).value != 'Y'):
                    fille = PatternFill("solid", fgColor="00FF00")
                    self.__m_sheet.cell(_index, _const.OSS_INFLUENCE).fill = fille
                self.__m_sheet.cell(_index, _const.OSS_INFLUENCE).value = searchinfo.get('influence')
                break


    def setData2Excel(self, _index, _tb):
        _t = TaskBean()
        _t = _tb
        self.__m_sheet.cell(_index,_const.PROJECT).value = _t.getProject()
        self.__m_sheet.cell(_index,_const.ISSUE_TYPE).value = _t.getIssuetype()
        self.__m_sheet.cell(_index, _const.SUMMARY).value = _t.getSummary()
        self.__m_sheet.cell(_index,_const.TEAMDEVELOPMENT).value = _t.getteamdevelopment()
        self.__m_sheet.cell(_index,_const.FUNCTIONALITY).value = _t.getFunctionality()
        self.__m_sheet.cell(_index,_const.ASSIGNEE).value = _t.getAssignee()
        self.__m_sheet.cell(_index,_const.TEAMTEST).value = _t.getteamtest()
        self.__m_sheet.cell(_index,_const.BUGFOUNDBY).value = _t.getbugfoundby()
        self.__m_sheet.cell(_index,_const.DESCRIPTION).value = _t.getDescription()
        self.__m_sheet.cell(_index,_const.REPRODUCIBILITY).value = _t.getreproducibility()
        self.__m_sheet.cell(_index,_const.SEVERITY).value = _t.getSeverity()
        self.__m_sheet.cell(_index,_const.MANUFACTURERVARIANT).value = _t.getManufacturervariant()
        self.__m_sheet.cell(_index,_const.FOUNDINSW).value = _t.getfoundinSW()
        self.__m_sheet.cell(_index, _const.FOUNDINHW).value = _t.getfoundinHW()
        self.__m_sheet.cell(_index, _const.FOUNDINHW).value = _t.getfoundinHW()
        self.__m_sheet.cell(_index, _const.LABELS).value = _t.getLabels()
        self.__m_sheet.cell(_index, _const.KEY).value = _t.getkey()
        self.__m_sheet.cell(_index, _const.COMMENTS).value = _t.getcomments()

    def saveExcel(self):
        self.__m_workBook.save(self.__m_filePath)

#de = CDoExcel('C:\\ZT\\Working\\紧急\\WBS\\WBS-OnlineApp-SOP2_V3_FieldsTemplate.xlsx')
#de.setSheetName('Import')
#max_row = de.getMaxrow()
#if max_row is not None:
#    for i in range(2, max_row+1):
#        issue_dict = de.getDataDict(i)
#        #print(issue_dict)
